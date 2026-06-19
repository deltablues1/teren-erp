"""Uvoz obračunskih projekata u "po sekciji" formatu.

Datoteka ima list po sekciji (1. RAZVODNI ORMARI … 13. REKAPITULACIJA), a svaki
ima isto zaglavlje: R.br. | Opis stavke | JM | Kol. | JC [€] | UC [€].

- UGOVORNI file: stupac D = UGOVORENA količina.
- Datoteka situacije: stupac D = KUMULATIVNA izvedena količina do te situacije,
  a stupac G = ugovorena količina.

Parser je deterministički (čiste kolone) — ne troši AI. Žargon/ključne riječi
ostaju prazni; popunjava ih kasnije kataloško prepoznavanje.

Read-only dio (parse_*) ništa ne piše u bazu. Pokreni `py -m services.situacija_import
"<putanja.xlsx>"` za dry-run pregled s kontrolom protiv Rekapitulacije.
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl

# Listovi koji NISU stavke troškovnika (match po ključnoj riječi u nazivu).
SKIP_KEYWORDS = ("naslovn", "opći uvjet", "opci uvjet", "opće napomen", "opce napomen",
                 "opći uvjeti", "rekapitulacij")
_KOMPLET_JM = {"kpl", "kompl", "kompl.", "komplet", "kpl."}


def _skip_sheet(title: str) -> bool:
    t = title.lower()
    return any(k in t for k in SKIP_KEYWORDS)


@dataclass
class Stavka:
    sifra: str
    sekcija: str
    pozicija: str
    opis: str
    jm: str
    ugovorena_kolicina: float | None
    jedinicna_cijena: float | None
    tip: str

    def as_row(self) -> list[Any]:
        """Redoslijed kao replace_troskovnik (11 kolona)."""
        return [
            self.sifra, self.sekcija, self.pozicija, self.opis, self.jm,
            "" if self.ugovorena_kolicina is None else self.ugovorena_kolicina,
            "" if self.jedinicna_cijena is None else self.jedinicna_cijena,
            self.tip, "", 0, "",
        ]


def _s(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _num(v: Any) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        f = float(v)
        return None if f != f else f
    try:
        return float(str(v).replace(",", ".").strip())
    except (TypeError, ValueError):
        return None


def _sekcija_broj(naziv: str) -> str:
    """'5. KABELI, POLICE I CIJEVI' → '5'."""
    m = re.match(r"\s*(\d+)", naziv)
    return m.group(1) if m else ""


def _header_redak(rows: list[tuple]) -> int | None:
    """Indeks retka sa zaglavljem. Prepoznaje per-sekcija zaglavlje ('Opis stavke')
    i plosnato/konsolidirano zaglavlje ('TROŠ. KOMAD' / 'PONUDA UKUPNO' — npr.
    list KABEL LISTA CARRIER)."""
    for idx, r in enumerate(rows):
        joined = " ".join(str(c).lower() for c in r if c)
        if "opis stavke" in joined or "troš. komad" in joined or "ponuda ukupno" in joined:
            return idx
    return None


def _iter_stavke(path: str | Path):
    """Generator preko svih sekcijskih listova. Yielda dict po stavci:
    {sifra, sekcija, pozicija, opis, jm, d, e, g} — gdje je
    D=stupac količine (ugovoreno u UGOVORNOM / kumulativ izvedenog u situaciji),
    E=jed. cijena, G=ugovorena količina (postoji u datotekama situacije)."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sec_ord = 0
    try:
        for ws in wb.worksheets:
            if _skip_sheet(ws.title):
                continue
            rows = list(ws.iter_rows(values_only=True))
            hdr = _header_redak(rows)
            if hdr is None:
                continue
            sec_ord += 1
            # Sekcijski broj: vodeća znamenka iz naziva ('5. KABELI'→'5') ili
            # redni broj lista ('RAZDJELNICI'→'S1') — da šifra ostane jedinstvena.
            secnum = _sekcija_broj(ws.title) or f"S{sec_ord}"
            sekcija = ws.title.strip()
            pozicija = ""
            for r in rows[hdr + 1:]:
                r = list(r) + [None] * (7 - len(r))
                a, b, c = _s(r[0]), _s(r[1]), _s(r[2])
                d, e, g = _num(r[3]), _num(r[4]), _num(r[6])
                if not b:
                    continue
                if not a or not a[0].isdigit():
                    continue  # napomena / nenumerirani tekst
                if not c:
                    pozicija = b  # naslov pozicije (nema JM)
                    continue
                sifra = f"{secnum}.{a.rstrip('.')}" if secnum else a.rstrip(".")
                yield {"sifra": sifra, "sekcija": sekcija, "pozicija": pozicija,
                       "opis": b, "jm": c, "d": d, "e": e, "g": g}
    finally:
        wb.close()


def parse_troskovnik(path: str | Path) -> list[Stavka]:
    """Pročitaj UGOVORNI file (D = ugovorena količina) → lista stavki."""
    stavke: list[Stavka] = []
    for it in _iter_stavke(path):
        if it["d"] is None:
            continue
        tip = "komplet" if it["jm"].lower() in _KOMPLET_JM else "stavka"
        stavke.append(Stavka(
            it["sifra"], it["sekcija"], it["pozicija"], it["opis"], it["jm"],
            it["d"], it["e"], tip,
        ))
    return stavke


def parse_troskovnik_flat(path: str | Path, sekcija: str = "Troškovnik") -> list[Stavka]:
    """Plosnati troškovnik (jedan list, bez sekcija): A=R.br, B=opis, C=JM,
    D=količina, E=cijena. Za jednostavne jednolistne ponude / ključ-u-ruke."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    stavke: list[Stavka] = []
    try:
        ws = max(wb.worksheets, key=lambda w: (w.max_row or 0) * (w.max_column or 0))
        rows = list(ws.iter_rows(values_only=True))
        hdr = None
        for i, r in enumerate(rows):
            joined = " ".join(str(c).lower() for c in r if c)
            if "komad" in joined or "ponuda ukupno" in joined or "jed. mjere" in joined:
                hdr = i
                break
        start = (hdr + 1) if hdr is not None else 0
        for r in rows[start:]:
            r = list(r) + [None] * (5 - len(r))
            a, b, c = _s(r[0]), _s(r[1]), _s(r[2])
            d, e = _num(r[3]), _num(r[4])
            if not b or not a or not a[0].isdigit():
                continue
            if not c or d is None:
                continue  # preskoči naslove/zbrojeve
            tip = "komplet" if c.lower() in _KOMPLET_JM else "stavka"
            stavke.append(Stavka(a.rstrip("."), sekcija, "", b, c, d, e, tip))
    finally:
        wb.close()
    return stavke


def parse_situacija(path: str | Path) -> dict[str, dict[str, Any]]:
    """Pročitaj datoteku situacije (D = kumulativna izvedena količina) →
    {sifra: {opis, jm, kumulativ}}. Stavke s kumulativom 0/praznim se preskoču."""
    out: dict[str, dict[str, Any]] = {}
    for it in _iter_stavke(path):
        kum = it["d"]
        if kum is None or kum == 0:
            continue
        out[it["sifra"]] = {"opis": it["opis"], "jm": it["jm"], "kumulativ": kum}
    return out


def import_troskovnik_u_bazu(
    path: str | Path,
    key: str,
    naziv: str,
    *,
    adresa: str = "",
    investitor: str = "",
    izvodac: str = "",
    nadzorni: str = "",
    flat: bool = False,
) -> dict[str, Any]:
    """Kreiraj projekt (ako ne postoji) i upiši troškovnik. Vrati sažetak.

    flat=True za plosnati jednolistni format, inače per-sekcija.
    Koristi db_backend izravno (Postgres). replace_troskovnik zamjenjuje sav
    postojeći sadržaj troškovnika tog projekta (idempotentno za ponovni uvoz)."""
    from services import db_backend

    stavke = parse_troskovnik_flat(path) if flat else parse_troskovnik(path)
    if not stavke:
        raise RuntimeError(f"Nijedna stavka nije izvučena iz {Path(path).name}.")

    novi = False
    try:
        db_backend.create_projekt(
            key, naziv, adresa=adresa, investitor=investitor,
            izvodac=izvodac, nadzorni=nadzorni,
        )
        novi = True
    except ValueError:
        pass  # projekt već postoji — samo zamijeni troškovnik

    n = db_backend.replace_troskovnik(key, [s.as_row() for s in stavke])
    vrijednost = sum(
        (s.ugovorena_kolicina or 0) * (s.jedinicna_cijena or 0) for s in stavke
    )
    return {"key": key, "projekt_kreiran": novi, "stavki": n,
            "ugovorena_vrijednost": round(vrijednost, 2)}


def import_situacija_u_bazu(
    path: str | Path,
    key: str,
    broj: int,
    *,
    datum: Any = None,
    status: str = "ovjerena",
) -> dict[str, Any]:
    """Uvezi datoteku situacije: snimi kumulativnu izvedenu količinu po stavci,
    matchiraj na troškovnik po šifri, ažuriraj izvedeno i izračunaj iznos ove
    situacije (kumulativ_ove − kumulativ_prethodne) × cijena.

    Idempotentno: postojeća situacija istog broja se zamijeni."""
    from sqlalchemy import delete, select

    from services import db
    from services.models import Situacija, SituacijaStavka, TroskovnikStavka

    sit_data = parse_situacija(path)
    with db.session() as s:
        trrows = s.scalars(
            select(TroskovnikStavka).where(TroskovnikStavka.projekt_key == key)
        ).all()
        by_sifra = {t.sifra: t for t in trrows}

        prev = s.scalar(
            select(Situacija)
            .where(Situacija.projekt_key == key, Situacija.broj < broj)
            .order_by(Situacija.broj.desc())
        )
        prev_kum: dict[str, float] = {}
        if prev:
            for ss in s.scalars(
                select(SituacijaStavka).where(SituacijaStavka.situacija_id == prev.id)
            ).all():
                prev_kum[ss.sifra] = ss.kolicina_kumulativ

        ex = s.scalar(
            select(Situacija).where(Situacija.projekt_key == key, Situacija.broj == broj)
        )
        if ex:
            s.execute(delete(SituacijaStavka).where(SituacijaStavka.situacija_id == ex.id))
            s.delete(ex)
            s.flush()

        sit = Situacija(projekt_key=key, broj=broj, datum=datum, status=status,
                        izvor_datoteka=Path(path).name)
        s.add(sit)
        s.flush()

        matched = 0
        nematchirano: list[str] = []
        kum_val = delta_val = 0.0
        for sifra, info in sit_data.items():
            t = by_sifra.get(sifra)
            cijena = (t.jedinicna_cijena if t and t.jedinicna_cijena else 0.0) or 0.0
            kum = info["kumulativ"]
            s.add(SituacijaStavka(
                situacija_id=sit.id, troskovnik_stavka_id=(t.id if t else None),
                sifra=sifra, opis=info["opis"], jm=info["jm"], kolicina_kumulativ=kum,
            ))
            if t:
                matched += 1
                t.izvedeno = kum
                kum_val += kum * cijena
                delta_val += (kum - prev_kum.get(sifra, 0.0)) * cijena
            else:
                nematchirano.append(sifra)

        return {
            "broj": broj, "stavki_u_situaciji": len(sit_data),
            "matchirano": matched, "nematchirano": nematchirano,
            "kumulativ_vrijednost": round(kum_val, 2),
            "iznos_ove_situacije": round(delta_val, 2),
        }


def _rekapitulacija(path: str | Path) -> dict[str, float]:
    """Pročitaj službene iznose iz lista Rekapitulacija (za kontrolu)."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    out: dict[str, float] = {}
    try:
        for ws in wb.worksheets:
            if "rekapitulacija" not in ws.title.lower():
                continue
            for r in ws.iter_rows(values_only=True):
                b = _s(r[1]) if len(r) > 1 else ""
                f = _num(r[5]) if len(r) > 5 else None
                if b and f is not None and any(
                    k in b.upper() for k in ("SVEUKUPNO", "UKUPNO", "RABAT")
                ):
                    out[b] = f
    finally:
        wb.close()
    return out


def _dry_run(path: str) -> None:
    import sys
    sys.stdout.reconfigure(encoding="utf-8")
    stavke = parse_troskovnik(path)
    print(f"Datoteka: {Path(path).name}")
    print(f"Izvučeno stavki: {len(stavke)}\n")

    po_sekciji: dict[str, list[Stavka]] = {}
    for s in stavke:
        po_sekciji.setdefault(s.sekcija, []).append(s)
    ukupno = 0.0
    print(f"{'SEKCIJA':<42} {'STAVKI':>7} {'VRIJEDNOST €':>14}")
    for sek, lst in po_sekciji.items():
        v = sum((s.ugovorena_kolicina or 0) * (s.jedinicna_cijena or 0) for s in lst)
        ukupno += v
        print(f"{sek[:42]:<42} {len(lst):>7} {v:>14,.2f}")
    print(f"{'— UKUPNO (Σ stavke)':<42} {len(stavke):>7} {ukupno:>14,.2f}")

    print("\nKONTROLA protiv Rekapitulacije iz datoteke:")
    for k, v in _rekapitulacija(path).items():
        print(f"  {k[:48]:<48} {v:>14,.2f}")

    print("\nPRIMJERI stavki (sifra | jm | ugovoreno | cijena | opis):")
    for s in stavke[:15]:
        print(f"  {s.sifra:<8} {s.jm:<6} {str(s.ugovorena_kolicina):>7} "
              f"{str(s.jedinicna_cijena):>7}  {s.opis[:55]}")


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Uporaba: py -m services.situacija_import \"<putanja troškovnika.xlsx>\"")
        sys.exit(1)
    _dry_run(sys.argv[1])
