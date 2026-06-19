"""Izvoz projekta u .xlsx — rekonstrukcija nekadašnjeg Google Sheets dokumenta
(„Teren - <projekt>.xlsx") iz Postgres baze.

5 listova kao u staroj Sheets eri (vidi services/sheets.py): Troskovnik, Dnevnik
(s originalnom porukom terenca u stupcu „Sirova_poruka"), Materijali, Radnici,
Vrijeme. Materijali su obogaćeni vezom na troškovničku poziciju i vrijednošću,
da dokument može služiti kao temelj za dnevnik i knjigu.
"""
from __future__ import annotations

import logging
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select

from config import GENERATED_DIR
from services import db, db_backend
from services.models import (
    DnevnikUnos,
    Materijal,
    Radnik,
    ProjektRadnik,
    Projekt,
    TroskovnikStavka,
    Vrijeme,
)

log = logging.getLogger(__name__)

_HDR_FILL = PatternFill("solid", fgColor="1F4E78")
_HDR_FONT = Font(bold=True, color="FFFFFF")


def _fmt_date(v: Any) -> str:
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    return "" if v is None else str(v)


def _write_sheet(ws, headers: list[str], rows: list[list[Any]]) -> None:
    """Upiši zaglavlje + retke, ostili header i postavi širine stupaca."""
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = Alignment(vertical="center")
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    # širine: po duljini sadržaja (ograničeno), s minimumom za čitljivost
    for i, _ in enumerate(headers, start=1):
        col = get_column_letter(i)
        maxlen = len(str(headers[i - 1]))
        for row in rows:
            if i - 1 < len(row) and row[i - 1] is not None:
                maxlen = max(maxlen, len(str(row[i - 1])))
        ws.column_dimensions[col].width = min(max(maxlen + 2, 10), 60)


def export_projekt(projekt_key: str) -> Path:
    """Složi .xlsx s 5 listova iz baze. Vrati putanju datoteke."""
    with db.session() as s:
        projekt = s.get(Projekt, projekt_key)
        if not projekt:
            raise ValueError(f"Projekt '{projekt_key}' ne postoji.")
        naziv = projekt.naziv or projekt_key

        # --- Troškovnik (s efektivnim izvedenim po stavci) ---
        eff = db_backend.izvedeno_efektivno(projekt_key)
        trosk = s.scalars(
            select(TroskovnikStavka)
            .where(TroskovnikStavka.projekt_key == projekt_key)
            .order_by(TroskovnikStavka.redoslijed, TroskovnikStavka.id)
        ).all()
        trosk_rows = []
        trosk_label: dict[int, str] = {}
        for st in trosk:
            izv = eff.get(st.id, st.izvedeno or 0.0)
            ugov = st.ugovorena_kolicina
            razlika = (ugov - izv) if (ugov is not None and st.tip != "sekcija") else None
            trosk_rows.append([
                st.sifra, st.sekcija, st.pozicija, st.opis, st.jm,
                ugov, st.jedinicna_cijena, st.tip, st.kljucne_rijeci,
                round(izv, 3) if izv else 0.0,
                round(razlika, 3) if razlika is not None else "",
            ])
            opis_kratko = (st.opis or "").strip()[:50]
            trosk_label[st.id] = f"{st.sifra} · {opis_kratko}".strip(" ·")

        # --- Dnevnik (s originalnom porukom) ---
        dnevnik = s.scalars(
            select(DnevnikUnos)
            .where(DnevnikUnos.projekt_key == projekt_key)
            .order_by(DnevnikUnos.datum, DnevnikUnos.upisano_at, DnevnikUnos.id)
        ).all()
        dnevnik_rows = [
            [
                _fmt_date(d.datum), _fmt_date(d.upisano_at), d.radnik, d.telegram_id,
                d.opis, d.lokacija, d.strujni_krug, d.vrijeme_rada,
                "" if d.sati is None else d.sati,
                d.radnici_spomenuti, d.problemi, d.sirova, d.confidence,
                d.telegram_msg_id,
            ]
            for d in dnevnik
        ]

        # --- Materijali (obogaćeni vezom na troškovnik + vrijednost) ---
        materijali = s.scalars(
            select(Materijal)
            .where(Materijal.projekt_key == projekt_key)
            .order_by(Materijal.datum, Materijal.id)
        ).all()
        cijena_po_id = {st.id: st.jedinicna_cijena for st in trosk}
        materijal_rows = []
        for m in materijali:
            tsid = m.troskovnik_stavka_id
            cijena = cijena_po_id.get(tsid) if tsid is not None else None
            vrijednost = round(cijena * m.kolicina, 2) if (cijena is not None and m.kolicina) else ""
            materijal_rows.append([
                _fmt_date(m.datum), m.vrijeme, m.radnik, m.telegram_id,
                m.sifra_stavke, m.opis, m.kolicina, m.jm, m.lokacija, m.napomena,
                m.strujni_krug,
                trosk_label.get(tsid, "") if tsid is not None else "",
                cijena if cijena is not None else "",
                vrijednost,
            ])

        # --- Radnici na projektu ---
        radnici = s.scalars(
            select(Radnik)
            .join(ProjektRadnik, ProjektRadnik.telegram_id == Radnik.telegram_id)
            .where(ProjektRadnik.projekt_key == projekt_key)
            .order_by(Radnik.ime)
        ).all()
        radnik_rows = [
            [r.telegram_id, r.ime, r.kvalifikacija, "Da" if r.aktivan else "Ne"]
            for r in radnici
        ]

        # --- Vrijeme ---
        vrijeme = s.scalars(
            select(Vrijeme)
            .where(Vrijeme.projekt_key == projekt_key)
            .order_by(Vrijeme.datum)
        ).all()
        vrijeme_rows = [
            [_fmt_date(v.datum), v.min_temp, v.max_temp,
             0 if v.oborine is None else v.oborine, v.opis]
            for v in vrijeme
        ]

    wb = Workbook()
    wb.remove(wb.active)  # makni defaultni prazni list

    _write_sheet(wb.create_sheet("Troskovnik"), [
        "Šifra", "Sekcija", "Pozicija", "Opis stavke", "JM",
        "Ugovorena količina", "Jedinična cijena", "Tip", "Ključne riječi",
        "Izvedeno", "Razlika",
    ], trosk_rows)

    _write_sheet(wb.create_sheet("Dnevnik"), [
        "Datum", "Upisano_at", "Radnik", "Telegram_ID", "Opis rada", "Lokacija",
        "Strujni_krug", "Vrijeme_rada", "Sati", "Radnici_spomenuti", "Problemi",
        "Sirova_poruka", "Confidence", "Telegram_msg_id",
    ], dnevnik_rows)

    _write_sheet(wb.create_sheet("Materijali"), [
        "Datum", "Vrijeme", "Radnik", "Telegram_ID", "Šifra_stavke", "Opis",
        "Količina", "JM", "Lokacija", "Napomena", "Strujni_krug",
        "Poz. troškovnika", "Jed. cijena", "Vrijednost",
    ], materijal_rows)

    _write_sheet(wb.create_sheet("Radnici"), [
        "Telegram_ID", "Ime", "Kvalifikacija", "Aktivan",
    ], radnik_rows)

    _write_sheet(wb.create_sheet("Vrijeme"), [
        "Datum", "Min_temp", "Max_temp", "Oborine_mm", "Vrijeme_opis",
    ], vrijeme_rows)

    safe_naziv = naziv.replace(" ", "_").replace("/", "-")
    out_path = GENERATED_DIR / f"Teren_{safe_naziv}_{datetime.now():%Y-%m-%d}.xlsx"
    wb.save(out_path)
    log.info("Izvezen projekt %s u %s", projekt_key, out_path)
    return out_path
