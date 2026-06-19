"""Čitanje Excel troškovnika - podržava .xlsx (openpyxl) i .xls (xlrd)."""
from __future__ import annotations

from pathlib import Path
from typing import Any


def read_rows(path: str | Path) -> list[list[Any]]:
    """Pročitaj prvi (najveći) list Excela u listu redaka.
    Svaki redak je lista vrijednosti ćelija. Format detektira po magic bajtovima
    (pouzdanije od ekstenzije - Telegram zna ne sačuvati .xls/.xlsx point)."""
    path = Path(path)

    with open(path, "rb") as f:
        header = f.read(8)

    # .xlsx (Office Open XML) je ZIP arhiva
    if header[:4] == b"PK\x03\x04":
        return _read_xlsx(path)
    # .xls (BIFF) je OLE Compound File
    if header[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        return _read_xls(path)

    # fallback na ekstenziju ako magic bajtovi ne prepoznaju
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return _read_xlsx(path)
    if suffix == ".xls":
        return _read_xls(path)
    raise ValueError(
        f"Nepodržan ili neprepoznat format: {path.name} "
        f"(podržano .xls, .xlsx)"
    )


def read_sheets(path: str | Path) -> list[tuple[str, list[list[Any]]]]:
    """Pročitaj SVE listove Excela: [(naziv_lista, redci), ...].

    Troškovnici su često razdijeljeni po listovima (jedan list = jedna sekcija:
    energetika, vatrodojava, uzemljenje…). `read_rows` vraća samo najveći list,
    pa bi se ostale sekcije izgubile — zato za uvoz troškovnika čitamo sve.
    Preskačemo potpuno prazne listove."""
    path = Path(path)
    with open(path, "rb") as f:
        header = f.read(8)

    if header[:4] == b"PK\x03\x04":
        sheets = _read_xlsx_sheets(path)
    elif header[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        sheets = _read_xls_sheets(path)
    elif path.suffix.lower() == ".xlsx":
        sheets = _read_xlsx_sheets(path)
    elif path.suffix.lower() == ".xls":
        sheets = _read_xls_sheets(path)
    else:
        raise ValueError(
            f"Nepodržan ili neprepoznat format: {path.name} (podržano .xls, .xlsx)"
        )

    out = []
    for name, rows in sheets:
        if any(any(c != "" for c in r) for r in rows):
            out.append((name, rows))
    return out


def _read_xlsx(path: Path) -> list[list[Any]]:
    """Učitaj preko BytesIO da openpyxl preskoči ekstenzijsku validaciju."""
    import openpyxl
    from io import BytesIO
    with open(path, "rb") as f:
        buf = BytesIO(f.read())
    wb = openpyxl.load_workbook(buf, data_only=True)
    ws = _largest_sheet_xlsx(wb)
    rows: list[list[Any]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append([_clean(v) for v in row])
    return rows


def _read_xlsx_sheets(path: Path) -> list[tuple[str, list[list[Any]]]]:
    import openpyxl
    from io import BytesIO
    with open(path, "rb") as f:
        buf = BytesIO(f.read())
    wb = openpyxl.load_workbook(buf, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        rows = [[_clean(v) for v in row] for row in ws.iter_rows(values_only=True)]
        sheets.append((ws.title, rows))
    return sheets


def _read_xls_sheets(path: Path) -> list[tuple[str, list[list[Any]]]]:
    import xlrd
    with open(path, "rb") as f:
        data = f.read()
    wb = xlrd.open_workbook(file_contents=data)
    sheets = []
    for sheet in wb.sheets():
        rows = [
            [_clean(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
            for r in range(sheet.nrows)
        ]
        sheets.append((sheet.name, rows))
    return sheets


def _largest_sheet_xlsx(wb):
    best = wb.active
    best_cells = 0
    for ws in wb.worksheets:
        cells = (ws.max_row or 0) * (ws.max_column or 0)
        if cells > best_cells:
            best_cells = cells
            best = ws
    return best


def _read_xls(path: Path) -> list[list[Any]]:
    """Učitaj preko file_contents da xlrd ne provjerava ime/ekstenziju."""
    import xlrd
    with open(path, "rb") as f:
        data = f.read()
    wb = xlrd.open_workbook(file_contents=data)
    sheet = max(wb.sheets(), key=lambda s: s.nrows * s.ncols)
    rows: list[list[Any]] = []
    for r in range(sheet.nrows):
        rows.append([_clean(sheet.cell_value(r, c)) for c in range(sheet.ncols)])
    return rows


def _clean(v: Any) -> Any:
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v
